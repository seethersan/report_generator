import unittest
import datetime
from openpyxl import load_workbook, Workbook

from report_generator import read_data, generate_report

class TestCreatingReport(unittest.TestCase):

    def test_read_data(self):
        """
        Tests that the read_data() function reads the expected data from the 
        'Analytics Template for Exercise.xlsx' file, and returns the correct
        sites_stats dictionary and stats_cols list.
        """
        expected_sites_stats = {
            'site 1': {
                datetime.datetime(2021, 1, 1): 
                {
                    'Page Views': 6,
                    'Unique Visitors': 4,
                    'Total Time Spent': 11,
                    "Visits": 4,
                    "Average Time Spent on Site": 0.1
                }
            },
            'site 2': {
                datetime.datetime(2021, 1, 1): 
                {
                    'Page Views': 4,
                    'Unique Visitors': 2,
                    'Total Time Spent': 5,
                    "Visits": 2,
                    "Average Time Spent on Site": 0
                }
            }
        }
                
        expected_stats_cols = ['Page Views', 'Unique Visitors', 'Total Time Spent', 'Visits', 'Average Time Spent on Site']
        
        sites_stats, stats_cols = read_data('test_file.xlsx')
        self.assertDictEqual(sites_stats, expected_sites_stats)
        self.assertListEqual(sorted(stats_cols), sorted(expected_stats_cols))
    
    def test_generate_report(self):
        """
        Tests that the generate_report() function creates a 'Result.xlsx' file
        that contains the contents of the expected_sites_stats dictionary.
        """
        expected_sites_stats = {
            'site 1': {
                datetime.date(2021, 1, 1): 
                {
                    'Page Views': 6,
                    'Unique Visitors': 4,
                    'Total Time Spent': 11,
                    "Visits": 4,
                    "Average Time Spent on Site": 0.1
                }
            },
            'site 2': {
                datetime.date(2021, 1, 1): 
                {
                    'Page Views': 4,
                    'Unique Visitors': 2,
                    'Total Time Spent': 5,
                    "Visits": 2,
                    "Average Time Spent on Site": 0
                }
            }
        }
                
        expected_stats_cols = ['Page Views', 'Unique Visitors', 'Total Time Spent', 'Visits', 'Average Time Spent on Site']
        
        generate_report('test_result.xlsx', expected_sites_stats, expected_stats_cols)
        
        # Open 'Result.xlsx' to check contents
        wb = load_workbook('test_result.xlsx')
        sheet = wb.worksheets[0]
        
        self.assertEqual(sheet['A1'].value, 'Day of Month')
        self.assertEqual(sheet['B1'].value, 'Date')
        self.assertEqual(sheet['C1'].value, 'Site ID')

        init_row = 2
        for site in expected_sites_stats:
            for date in expected_sites_stats[site]:
                self.assertEqual(sheet.cell(row=init_row, column=1).value, date.day)
                self.assertEqual(sheet.cell(row=init_row, column=2).value, date.strftime("%Y/%m/%d"))
                self.assertEqual(sheet.cell(row=init_row, column=3).value, site)
            
                for i in range(len(expected_stats_cols)):
                    self.assertEqual(sheet.cell(row=init_row, column=i + 4).value, expected_sites_stats[site][date][expected_stats_cols[i]])
            
                init_row += 1

if __name__ == '__main__':
    unittest.main()
