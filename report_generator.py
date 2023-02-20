from openpyxl import load_workbook, Workbook

def read_data(file_name):
    """
    Reads data from provided file and returns the extracted data.

    Parameters: 
    file_name (string): The path of the Excel file to read.

    Returns:
    sites_stats (list): A list containing information about each site.
    stats_cols (list): A list containing all of the stats in the specified file.
    """
    wb = load_workbook(file_name)
    sheet = wb.worksheets[0]

    start_date = sheet['A1'].value
    end_date = sheet['A2'].value

    sites_stats = {}
    stats_cols = set([])

    for i in range(4, sheet.max_row):
        site = sheet.cell(row=i, column=1).value
        if site is None:
            continue
        sites_stats[site] = {}
        for j in range(2, sheet.max_column + 1):
            stat_cell = sheet.cell(row=i - 2, column=j)
            date_cell = sheet.cell(row=i - 1, column=j)
            stats_cols.add(stat_cell.value)
            if date_cell.value < start_date or date_cell.value > end_date:
                continue
            if not sites_stats[site].get(date_cell.value):
                sites_stats[site][date_cell.value] = {
                    stat_cell.value: sheet.cell(row=i, column=j).value
                }
            else:
                sites_stats[site][date_cell.value][stat_cell.value] = sheet.cell(row=i, column=j).value
    return sites_stats, list(stats_cols)

def generate_report(file_name, sites_stats, stats_cols):
    """
    This function generates a report based on the given Excel file, sites_stats and stats_cols.
    It writes data from the sites_stats dictionary into an newly created Excel file,
    adding the contents of stats_cols to the first row.
    """
    wb = Workbook()
    sheet = wb.worksheets[0]

    sheet["A1"] = "Day of Month"
    sheet["B1"] = "Date"
    sheet["C1"] = "Site ID"

    for i in range(len(stats_cols)):
        sheet.cell(row=1, column=i + 4).value = stats_cols[i]

    init_row = 2

    for site in sites_stats:
        for date in sites_stats[site]:
            sheet.cell(row=init_row, column=1).value = date.day
            sheet.cell(row=init_row, column=2).value = date.strftime("%Y/%m/%d")
            sheet.cell(row=init_row, column=3).value = site
            for i in range(len(stats_cols)):
                sheet.cell(row=init_row, column=i + 4).value = sites_stats[site][date].get(stats_cols[i], '-')
            init_row += 1

    wb.save(file_name)

if __name__ == "__main__":
    sites_stats, stats_cols = read_data('Analytics Template for Exercise.xlsx')
    generate_report('Result.xlsx', sites_stats, stats_cols)